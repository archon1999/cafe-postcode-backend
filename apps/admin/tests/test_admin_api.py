from datetime import UTC, datetime, timedelta
from io import BytesIO
from unittest.mock import patch

from django.core.exceptions import RequestDataTooBig
from django.urls import resolve
from django.utils import timezone, translation
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase

from apps.accounts.models import AuthSession, Permission, Role, User
from apps.floor.models import DiningTable, Hall, TableSession, ZoneOrCabin
from apps.integrations.models import IntegrationConfig
from apps.orders.models import Order, Payment
from apps.organizations.models import BusinessPartner, Restaurant, RestaurantEntitlement
from apps.organizations.services.faktura import FakturaError
from common.api.exception_handler import custom_exception_handler
from common.utils.date import TASHKENT_TIMEZONE


class AdminApiTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.restaurant = Restaurant.objects.create(
            name='Test Restaurant',
            phone='+998900000000',
            address='Tashkent',
        )

        permission_codes = [
            'permissions.list',
            'roles.list',
            'users.list',
            'reports.summary.view',
            'reports.shifts.view',
            'reports.shifts.export',
        ]
        cls.permissions = {
            code: Permission.objects.get_or_create(
                code=code,
                defaults={'name': code, 'description': f'{code} permission'},
            )[0]
            for code in permission_codes
        }

        cls.admin_role, _ = Role.objects.get_or_create(
            code='admin-manager',
            defaults={
                'name': 'Admin Manager',
                'description': 'Admin role for tests',
                'is_system': False,
            },
        )
        cls.admin_role.permissions.set(cls.permissions.values())
        cls.entitlement = RestaurantEntitlement.objects.create(
            restaurant=cls.restaurant,
            is_active=True,
            is_custom=True,
        )
        cls.default_zone = ZoneOrCabin.objects.create(
            restaurant=cls.restaurant,
            name='Default zone',
            sort_order=1,
            is_active=True,
        )
        cls.entitlement.permissions.set(cls.permissions.values())
        cls.entitlement.allowed_roles.set([cls.admin_role, cls.limited_role] if hasattr(cls, 'limited_role') else [cls.admin_role])

        cls.admin_user = User.objects.create_user(
            username='admin-user',
            password='secret123',
            full_name='Admin User',
            restaurant=cls.restaurant,
            role=cls.admin_role,
            actor_type=User.ActorType.RESTAURANT_ADMIN,
            ui_mode=User.UiMode.ADMIN,
            is_staff=True,
        )

        cls.superuser = User.objects.create_superuser(
            username='super-admin',
            password='secret123',
            full_name='Super Admin',
        )

        cls.limited_role, _ = Role.objects.get_or_create(
            code='limited-admin',
            defaults={
                'name': 'Limited Admin',
                'description': 'Limited permissions',
                'is_system': False,
            },
        )
        cls.limited_role.permissions.set([cls.permissions['users.list']])
        cls.entitlement.allowed_roles.add(cls.limited_role)

        cls.limited_user = User.objects.create_user(
            username='limited-user',
            password='secret123',
            full_name='Limited User',
            restaurant=cls.restaurant,
            role=cls.limited_role,
            actor_type=User.ActorType.RESTAURANT_ADMIN,
            ui_mode=User.UiMode.ADMIN,
            is_staff=True,
        )

    def authenticate(self, user=None):
        self.client.force_authenticate(user=user or self.admin_user)

    @staticmethod
    def _workbook_values(response):
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        return [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell not in (None, '')]

    def test_admin_routes_resolve_to_admin_views(self):
        match = resolve('/api/v1/admin/auth/login/')
        self.assertEqual(match.func.view_class.__module__, 'apps.admin.views.auth')

        match = resolve('/api/v1/admin/reports/summary/')
        self.assertEqual(match.func.view_class.__module__, 'apps.admin.views.reports')

    def test_admin_auth_login_me_logout_cycle(self):
        response = self.client.post(
            '/api/v1/admin/auth/login/',
            {'username': self.admin_user.username, 'password': 'secret123'},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('token', response.data)
        self.assertIn('session', response.data)
        self.assertEqual(response.data['user']['username'], self.admin_user.username)
        session_id = response.data['session']['id']

        token = response.data['token']
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token}')

        me_response = self.client.get('/api/v1/admin/auth/me/')
        self.assertEqual(me_response.status_code, status.HTTP_200_OK)
        self.assertEqual(me_response.data['username'], self.admin_user.username)
        system_time = datetime.fromisoformat(me_response.headers['system-time'])
        self.assertEqual(system_time.utcoffset(), timedelta(hours=5))

        logout_response = self.client.post('/api/v1/admin/auth/logout/')
        self.assertEqual(logout_response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Token.objects.filter(user=self.admin_user).exists())
        session = AuthSession.objects.get(pk=session_id)
        self.assertEqual(session.status, AuthSession.Status.REVOKED)
        self.assertIsNotNone(session.revoked_at)

    def test_admin_permissions_are_enforced(self):
        self.authenticate(self.limited_user)

        response = self.client.get('/api/v1/admin/reports/summary/')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_users_roles_and_permissions_endpoints(self):
        self.authenticate()

        permissions_response = self.client.get('/api/v1/admin/users/permissions/')
        self.assertEqual(permissions_response.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(permissions_response.data['total'], len(self.permissions))

        permission_options_response = self.client.get('/api/v1/admin/users/permissions/options/')
        self.assertEqual(permission_options_response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(permission_options_response.data, list)
        self.assertGreaterEqual(len(permission_options_response.data), len(self.permissions))

        roles_response = self.client.get('/api/v1/admin/users/roles/')
        self.assertEqual(roles_response.status_code, status.HTTP_200_OK)
        self.assertEqual(roles_response.data['data'][0]['code'], self.admin_role.code)

        users_response = self.client.get('/api/v1/admin/users/')
        self.assertEqual(users_response.status_code, status.HTTP_200_OK)
        usernames = {row['username'] for row in users_response.data['data']}
        self.assertIn(self.admin_user.username, usernames)

    def test_admin_feature_config_can_be_updated(self):
        self.authenticate(self.superuser)

        response = self.client.patch(
            f'/api/v1/admin/restaurants/{self.restaurant.id}/feature-config/',
            {'owner_dashboard_enabled': False},
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data['owner_dashboard_enabled'])
        self.assertEqual(str(response.data['restaurant']), str(self.restaurant.id))

    def test_admin_feature_config_returns_enabled_role_details(self):
        self.authenticate(self.superuser)
        self.feature_config.enabled_roles = [self.admin_role.code]
        self.feature_config.save(update_fields=['enabled_roles'])

        response = self.client.get(f'/api/v1/admin/restaurants/{self.restaurant.id}/feature-config/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data['enabled_role_details'],
            [{'id': str(self.admin_role.id), 'code': self.admin_role.code, 'name': self.admin_role.name}],
        )

    def test_restaurant_feature_config_is_superuser_only(self):
        self.authenticate(self.admin_user)

        response = self.client.get(f'/api/v1/admin/restaurants/{self.restaurant.id}/feature-config/')

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_restaurant_create_works_without_slug(self):
        self.authenticate(self.superuser)

        response = self.client.post(
            '/api/v1/admin/constructor/restaurants/',
            {
                'name': 'Second Restaurant',
                'legal_name': 'Second Restaurant LLC',
                'tax_number': '123456',
                'phone': '+998900000010',
                'address': 'Samarkand',
                'is_active': True,
            },
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['name'], 'Second Restaurant')
        self.assertEqual(response.data['currency'], 'UZS')
        self.assertNotIn('timezone', response.data)

    def test_report_summary_uses_tashkent_day_boundaries(self):
        self.authenticate()

        report_dt = datetime(2026, 3, 26, 19, 30, tzinfo=UTC)
        hall = Hall.objects.create(
            zone_or_cabin=self.default_zone,
            name='Summary Hall',
        )
        table = DiningTable.objects.create(
            hall=hall,
            zone=None,
            name='B1',
            table_number=2,
            seat_count=4,
            shape=DiningTable.Shape.SQUARE,
            shape_variant=DiningTable.ShapeVariant.SEAT4_SQUARE,
            status=DiningTable.Status.OCCUPIED,
            position_x=0,
            position_y=0,
            width=1,
            height=1,
        )
        session = TableSession.objects.create(
            restaurant=self.restaurant,
            hall=hall,
            table=table,
            opened_by=self.admin_user,
            guest_count=2,
            status=TableSession.Status.OPEN,
        )
        order = Order.objects.create(
            restaurant=self.restaurant,
            table_session=session,
            opened_by=self.admin_user,
            order_number=102,
            status=Order.Status.CLOSED,
            subtotal=120_000,
            total=120_000,
            closed_at=report_dt,
        )
        payment = Payment.objects.create(
            order=order,
            received_by=self.admin_user,
            method=Payment.Method.CARD,
            amount=120_000,
            status=Payment.Status.SUCCEEDED,
            paid_at=report_dt,
        )
        Order.objects.filter(pk=order.pk).update(created_at=report_dt, updated_at=report_dt, closed_at=report_dt)
        TableSession.objects.filter(pk=session.pk).update(created_at=report_dt, updated_at=report_dt)
        Payment.objects.filter(pk=payment.pk).update(created_at=report_dt, updated_at=report_dt, paid_at=report_dt)

        response = self.client.get(
            '/api/v1/admin/reports/summary/',
            {'period_type': 'day', 'date': '2026-03-27'},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['sales_total'], 120000)
        self.assertEqual(response.data['orders_count'], 1)

    def test_superuser_report_summary_without_restaurant_scope_aggregates_all_restaurants(self):
        self.authenticate(self.superuser)

        second_restaurant = Restaurant.objects.create(
            name='Second Restaurant',
            phone='+998900000111',
            address='Samarkand',
        )

        def create_sale(restaurant, order_number, amount):
            zone = ZoneOrCabin.objects.create(
                restaurant=restaurant,
                name=f'Zone {order_number}',
                sort_order=order_number,
                is_active=True,
            )
            hall = Hall.objects.create(zone_or_cabin=zone, name=f'Hall {order_number}')
            table = DiningTable.objects.create(
                hall=hall,
                zone=None,
                name=f'Table {order_number}',
                table_number=order_number,
                seat_count=4,
                shape=DiningTable.Shape.SQUARE,
                shape_variant=DiningTable.ShapeVariant.SEAT4_SQUARE,
                status=DiningTable.Status.OCCUPIED,
                position_x=0,
                position_y=0,
                width=1,
                height=1,
            )
            session = TableSession.objects.create(
                restaurant=restaurant,
                hall=hall,
                table=table,
                opened_by=self.admin_user,
                guest_count=2,
                status=TableSession.Status.OPEN,
            )
            report_dt = timezone.now()
            order = Order.objects.create(
                restaurant=restaurant,
                table_session=session,
                opened_by=self.admin_user,
                order_number=order_number,
                status=Order.Status.CLOSED,
                subtotal=amount,
                total=amount,
                closed_at=report_dt,
            )
            Payment.objects.create(
                order=order,
                received_by=self.admin_user,
                method=Payment.Method.CASH,
                amount=amount,
                status=Payment.Status.SUCCEEDED,
                paid_at=report_dt,
            )

        create_sale(self.restaurant, 301, 120000)
        create_sale(second_restaurant, 302, 80000)

        response = self.client.get('/api/v1/admin/reports/summary/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['sales_total'], 200000)
        self.assertEqual(response.data['orders_count'], 2)

    def test_hall_create_works_without_branch_fields(self):
        self.authenticate()

        hall_response = self.client.post(
            '/api/v1/admin/floor/halls/',
            {
                'name': 'Blue Hall',
                'description': 'Test hall',
                'grid_columns': 8,
                'sort_order': 0,
                'is_active': True,
                'zoneOrCabinId': str(self.default_zone.id),
            },
            format='json',
        )
        self.assertEqual(hall_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(hall_response.data['name'], 'Blue Hall')

    def test_admin_catalog_category_create_and_update(self):
        self.authenticate()

        create_response = self.client.post(
            '/api/v1/admin/catalog/categories/',
            {
                'name': 'Hot Drinks',
                'mxik_code': '10000000000000001',
                'mxik_name': 'Hot Drinks',
                'name_uz': 'Issiq ichimliklar',
                'name_uz_crl': 'Иссиқ ичимликлар',
                'name_ru': 'Горячие напитки',
                'kind': 'drink',
                'sort_order': 1,
                'is_active': True,
            },
            format='json',
        )
        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        category_id = create_response.data['id']

        update_response = self.client.patch(
            f'/api/v1/admin/catalog/categories/{category_id}/',
            {'name': 'Hot Drinks Updated'},
            format='json',
        )
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        self.assertEqual(update_response.data['name'], 'Hot Drinks Updated')

    def test_catalog_create_work_without_code(self):
        self.authenticate()

        category_response = self.client.post(
            '/api/v1/admin/catalog/categories/',
            {
                'name': 'Cold Drinks',
                'mxik_code': '10000000000000002',
                'mxik_name': 'Cold Drinks',
                'name_uz': 'Sovuq ichimliklar',
                'name_uz_crl': 'Совуқ ичимликлар',
                'name_ru': 'Холодные напитки',
                'kind': 'drink',
                'sort_order': 2,
                'is_active': True,
            },
            format='json',
        )
        self.assertEqual(category_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(category_response.data['name'], 'Cold Drinks')

        item_response = self.client.post(
            '/api/v1/admin/catalog/items/',
            {
                'category': str(category_response.data['id']),
                'name': 'Iced Tea',
                'name_uz': 'Muzli choy',
                'name_uz_crl': 'Музли чой',
                'name_ru': 'Холодный чай',
                'kind': 'drink',
                'description': '',
                'description_uz': '',
                'description_uz_crl': '',
                'description_ru': '',
                'sku': '',
                'is_active': True,
                'is_stoplisted': False,
            },
            format='json',
        )
        self.assertEqual(item_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(item_response.data['name'], 'Iced Tea')

    def test_admin_catalog_category_returns_image_fields_and_items_support_active_filter(self):
        self.authenticate()

        category_response = self.client.post(
            '/api/v1/admin/catalog/categories/',
            {
                'name': 'Bakery',
                'mxik_code': '10000000000000003',
                'mxik_name': 'Bakery',
                'sort_order': 3,
                'is_active': True,
            },
            format='json',
        )
        self.assertEqual(category_response.status_code, status.HTTP_201_CREATED)
        self.assertIn('image_url', category_response.data)
        self.assertIn('image_source', category_response.data)

        active_item = self.client.post(
            '/api/v1/admin/catalog/items/',
            {
                'category': str(category_response.data['id']),
                'name': 'Croissant',
                'description': '',
                'is_active': True,
                'is_stoplisted': False,
            },
            format='json',
        )
        self.assertEqual(active_item.status_code, status.HTTP_201_CREATED)

        inactive_item = self.client.post(
            '/api/v1/admin/catalog/items/',
            {
                'category': str(category_response.data['id']),
                'name': 'Old Pie',
                'description': '',
                'is_active': False,
                'is_stoplisted': False,
            },
            format='json',
        )
        self.assertEqual(inactive_item.status_code, status.HTTP_201_CREATED)

        items_response = self.client.get('/api/v1/admin/catalog/items/', {'is_active': 'true'})
        self.assertEqual(items_response.status_code, status.HTTP_200_OK)
        item_names = {row['name'] for row in items_response.data['data']}
        self.assertIn('Croissant', item_names)
        self.assertNotIn('Old Pie', item_names)

    def test_admin_integration_config_create_and_update(self):
        self.authenticate()

        create_response = self.client.post(
            '/api/v1/admin/integrations/configs/',
            {
                'kind': IntegrationConfig.Kind.FISCAL,
                'provider': 'soliq-service',
                'mode': IntegrationConfig.Mode.MOCK,
                'is_enabled': True,
                'settings': {'terminalId': 'T-1'},
            },
            format='json',
        )
        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        config_id = create_response.data['id']

        update_response = self.client.patch(
            f'/api/v1/admin/integrations/configs/{config_id}/',
            {'provider': 'soliq-service-v2'},
            format='json',
        )
        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        self.assertEqual(update_response.data['provider'], 'soliq-service-v2')

    def test_admin_reports_summary_and_export(self):
        self.authenticate()

        summary_response = self.client.get(
            '/api/v1/admin/reports/summary/',
            {'period_type': 'day', 'date': '2026-03-27'},
        )
        self.assertEqual(summary_response.status_code, status.HTTP_200_OK)
        self.assertEqual(summary_response.data['sales_total'], 0)
        self.assertEqual(summary_response.data['orders_count'], 0)

        summary_export_response = self.client.get(
            '/api/v1/admin/reports/summary/export/',
            {'period_type': 'day', 'date': '2026-03-27'},
        )
        self.assertEqual(summary_export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            summary_export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('summary-report-day-2026-03-27.xlsx', summary_export_response['Content-Disposition'])

        sales_response = self.client.get(
            '/api/v1/admin/reports/sales/',
            {'period_type': 'month', 'month': '2026-03', 'page_size': 10},
        )
        self.assertEqual(sales_response.status_code, status.HTTP_200_OK)
        self.assertEqual(sales_response.data['pageSize'], 10)
        self.assertEqual(sales_response.data['total'], 0)
        self.assertEqual(sales_response.data['data'], [])

        export_response = self.client.get(
            '/api/v1/admin/reports/sales/export/',
            {'period_type': 'month', 'month': '2026-03'},
        )
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('sales-report-month-2026-03.xlsx', export_response['Content-Disposition'])

        open_checks_export_response = self.client.get(
            '/api/v1/admin/reports/open-checks/export/',
            {'period_type': 'year', 'year': '2026'},
        )
        self.assertEqual(open_checks_export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            open_checks_export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('open-checks-report-year-2026.xlsx', open_checks_export_response['Content-Disposition'])

        top_items_response = self.client.get(
            '/api/v1/admin/reports/top-items/',
            {'period_type': 'month', 'month': '2026-03', 'page_size': 10},
        )
        self.assertEqual(top_items_response.status_code, status.HTTP_200_OK)
        self.assertEqual(top_items_response.data['pageSize'], 10)
        self.assertEqual(top_items_response.data['total'], 0)
        self.assertEqual(top_items_response.data['data'], [])

        top_items_export_response = self.client.get(
            '/api/v1/admin/reports/top-items/export/',
            {'period_type': 'month', 'month': '2026-03'},
        )
        self.assertEqual(top_items_export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            top_items_export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('top-items-report-month-2026-03.xlsx', top_items_export_response['Content-Disposition'])

    def test_admin_report_exports_are_localized_by_request_language(self):
        self.authenticate()

        report_dt = datetime(2026, 3, 27, 10, 0, tzinfo=TASHKENT_TIMEZONE)
        hall = Hall.objects.create(
            zone_or_cabin=self.default_zone,
            name='Main Hall',
        )
        table = DiningTable.objects.create(
            hall=hall,
            zone=None,
            name='A1',
            table_number=1,
            seat_count=4,
            shape=DiningTable.Shape.SQUARE,
            shape_variant=DiningTable.ShapeVariant.SEAT4_SQUARE,
            status=DiningTable.Status.OCCUPIED,
            position_x=0,
            position_y=0,
            width=1,
            height=1,
        )
        session = TableSession.objects.create(
            restaurant=self.restaurant,
            hall=hall,
            table=table,
            opened_by=self.admin_user,
            guest_count=2,
            status=TableSession.Status.OPEN,
        )
        order = Order.objects.create(
            restaurant=self.restaurant,
            table_session=session,
            opened_by=self.admin_user,
            order_number=101,
            status=Order.Status.OPEN,
            subtotal=100_000,
            total=100_000,
        )
        payment = Payment.objects.create(
            order=order,
            received_by=self.admin_user,
            method=Payment.Method.CARD,
            amount=100_000,
            status=Payment.Status.SUCCEEDED,
            paid_at=report_dt,
        )
        Order.objects.filter(pk=order.pk).update(created_at=report_dt, updated_at=report_dt)
        TableSession.objects.filter(pk=session.pk).update(created_at=report_dt, updated_at=report_dt)
        Payment.objects.filter(pk=payment.pk).update(created_at=report_dt, updated_at=report_dt, paid_at=report_dt)

        sales_export_ru = self.client.get(
            '/api/v1/admin/reports/sales/export/',
            {'period_type': 'day', 'date': '2026-03-27', 'lang': 'ru'},
        )
        self.assertEqual(sales_export_ru.status_code, status.HTTP_200_OK)
        sales_values_ru = self._workbook_values(sales_export_ru)
        self.assertIn('Отчет по продажам', sales_values_ru)
        self.assertIn('Способ', sales_values_ru)
        self.assertIn('Карта', sales_values_ru)

        open_checks_export_uz_crl = self.client.get(
            '/api/v1/admin/reports/open-checks/export/',
            {'period_type': 'day', 'date': '2026-03-27', 'lang': 'uz-crl'},
        )
        self.assertEqual(open_checks_export_uz_crl.status_code, status.HTTP_200_OK)
        open_checks_values_uz_crl = self._workbook_values(open_checks_export_uz_crl)
        self.assertIn('Очиқ чеклар ҳисоботи', open_checks_values_uz_crl)
        self.assertIn('Ҳолат', open_checks_values_uz_crl)
        self.assertIn('Очиқ', open_checks_values_uz_crl)

        summary_export_fallback = self.client.get(
            '/api/v1/admin/reports/summary/export/',
            {'period_type': 'day', 'date': '2026-03-27', 'lang': 'de'},
        )
        self.assertEqual(summary_export_fallback.status_code, status.HTTP_200_OK)
        summary_values_fallback = self._workbook_values(summary_export_fallback)
        self.assertIn('Hisobot xulosasi', summary_values_fallback)

    def test_admin_errors_are_localized_by_request_language(self):
        login_response = self.client.post(
            '/api/v1/admin/auth/login/?lang=ru',
            {'username': self.admin_user.username, 'password': 'wrong-secret'},
            format='json',
        )
        self.assertEqual(login_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(login_response.data['non_field_errors'][0], 'Неверные учетные данные.')

        self.authenticate()
        create_user_response = self.client.post(
            '/api/v1/admin/users/?lang=uz-crl',
            {
                'username': 'pin-error-user',
                'full_name': 'Pin Error User',
                'ui_mode': User.UiMode.POS,
                'is_active': True,
                'role_id': str(self.admin_role.id),
                'pin': 123456,
            },
            format='json',
        )
        self.assertEqual(create_user_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(create_user_response.data['pin'][0], 'PIN-код сатр бўлиши керак.')

        with translation.override('uz-crl'):
            response = custom_exception_handler(RequestDataTooBig('too large'), {})
        self.assertEqual(response.status_code, status.HTTP_413_REQUEST_ENTITY_TOO_LARGE)
        self.assertEqual(response.data['message'], 'Юкланган файл жуда катта.')
    @patch('apps.admin.views.product_owner.FakturaClient.lookup_company_basic_details')
    def test_business_partner_lookup_returns_normalized_payload(self, lookup_mock):
        self.authenticate(self.superuser)
        lookup_mock.return_value = {
            'CompanyInn': '123456789',
            'CompanyName': 'Acme',
            'DirectorName': 'John Doe',
            'PhoneNumber': '+998901234567',
            'Email': 'info@acme.uz',
            'CompanyAddress': 'Tashkent',
        }

        response = self.client.get('/api/v1/admin/platform/business-partners/lookup/', {'inn': '123456789'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['inn'], '123456789')
        self.assertEqual(response.data['company_name'], 'Acme')
        self.assertEqual(response.data['legal_name'], 'Acme')
        self.assertEqual(response.data['director_name'], 'John Doe')
        self.assertEqual(response.data['phone'], '+998901234567')
        self.assertEqual(response.data['email'], 'info@acme.uz')
        self.assertEqual(response.data['address'], 'Tashkent')
        self.assertEqual(response.data['faktura_payload']['CompanyName'], 'Acme')

    def test_business_partner_lookup_requires_inn(self):
        self.authenticate(self.superuser)

        response = self.client.get('/api/v1/admin/platform/business-partners/lookup/')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('inn', response.data)

    @patch('apps.admin.views.product_owner.FakturaClient.lookup_company_basic_details')
    def test_business_partner_lookup_returns_502_on_upstream_failure(self, lookup_mock):
        self.authenticate(self.superuser)
        lookup_mock.side_effect = FakturaError('boom')

        response = self.client.get('/api/v1/admin/platform/business-partners/lookup/', {'inn': '123456789'})

        self.assertEqual(response.status_code, status.HTTP_502_BAD_GATEWAY)
        self.assertEqual(response.data['detail'], 'boom')

    def test_business_partner_lookup_is_permission_protected(self):
        self.authenticate(self.limited_user)

        response = self.client.get('/api/v1/admin/platform/business-partners/lookup/', {'inn': '123456789'})

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_business_partner_create_and_update_preserve_faktura_payload(self):
        self.authenticate(self.superuser)

        create_response = self.client.post(
            '/api/v1/admin/platform/business-partners/',
            {
                'inn': '123456789',
                'companyName': 'Acme',
                'legalName': 'Acme LLC',
                'directorName': 'John Doe',
                'phone': '+998901234567',
                'email': 'info@acme.uz',
                'address': 'Tashkent',
                'fakturaPayload': {'CompanyName': 'Acme', 'CompanyInn': '123456789'},
            },
            format='json',
        )

        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        partner = BusinessPartner.objects.get(pk=create_response.data['id'])
        self.assertEqual(partner.faktura_payload['CompanyName'], 'Acme')

        update_response = self.client.patch(
            f'/api/v1/admin/platform/business-partners/{partner.id}/',
            {
                'companyName': 'Acme Updated',
            },
            format='json',
        )

        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        partner.refresh_from_db()
        self.assertEqual(partner.company_name, 'Acme Updated')
        self.assertEqual(partner.faktura_payload['CompanyName'], 'Acme')
