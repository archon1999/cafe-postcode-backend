from datetime import timedelta
from io import BytesIO

from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from rest_framework import status

from apps.orders.models import Order, OrderItem, Payment, Receipt
from apps.orders.services import CashShiftService
from apps.reports.services.export_localization import get_open_checks_columns, get_sales_columns, get_shift_columns
from common.tests.pos_api import PosAPITestCase


class ReportsApiTests(PosAPITestCase):
    def setUp(self):
        super().setUp()
        now = timezone.now()
        self.shift_service = CashShiftService()
        self.closed_order = Order.objects.create(
            restaurant=self.restaurant,
            branch=self.branch,
            distribution_point=self.takeaway_distribution,
            opened_by=self.user,
            cashier=self.user,
            order_number=1,
            channel=Order.Channel.TAKEAWAY,
            status=Order.Status.CLOSED,
            guest_count=1,
            closed_at=now - timedelta(hours=1),
        )
        OrderItem.objects.create(
            order=self.closed_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=30000,
            status=OrderItem.Status.DONE,
        )
        self.closed_order.recalculate_totals()
        self.open_order = Order.objects.create(
            restaurant=self.restaurant,
            branch=self.branch,
            distribution_point=self.hall_distribution,
            opened_by=self.user,
            order_number=2,
            channel=Order.Channel.HALL,
            status=Order.Status.SUBMITTED,
            guest_count=2,
        )
        OrderItem.objects.create(
            order=self.open_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=30000,
            status=OrderItem.Status.NEW,
        )
        self.open_order.recalculate_totals()

        second_restaurant = self.restaurant.__class__.objects.create(name='Other restaurant')
        second_branch = self.branch.__class__.objects.create(
            restaurant=second_restaurant,
            name='Other branch',
            service_fee_percent=10,
            is_default=True,
        )
        other_order = Order.objects.create(
            restaurant=second_restaurant,
            branch=second_branch,
            distribution_point=self.takeaway_distribution.__class__.objects.create(
                restaurant=second_restaurant,
                branch=second_branch,
                name='Other takeaway',
                kind=self.takeaway_distribution.kind,
            ),
            opened_by=self.user,
            cashier=self.user,
            order_number=1,
            channel=Order.Channel.TAKEAWAY,
            status=Order.Status.CLOSED,
            guest_count=1,
            closed_at=now - timedelta(hours=1),
        )
        OrderItem.objects.create(
            order=other_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=50000,
            status=OrderItem.Status.DONE,
        )
        other_order.recalculate_totals()
        Payment.objects.create(
            order=other_order,
            received_by=self.user,
            method=Payment.Method.CASH,
            amount=other_order.total,
            status=Payment.Status.SUCCEEDED,
            paid_at=now - timedelta(hours=1),
        )

        self.shift = self.create_cash_shift(opening_cash_amount=100000)
        self.shift_payment = Payment.objects.create(
            order=self.closed_order,
            cash_shift=self.shift,
            cash_desk=self.cash_desk,
            received_by=self.user,
            method=Payment.Method.CASH,
            amount=self.closed_order.total,
            status=Payment.Status.SUCCEEDED,
            paid_at=now - timedelta(minutes=30),
        )
        Receipt.objects.create(
            order=self.closed_order,
            payment=self.shift_payment,
            kind=Receipt.Kind.FISCAL,
            status=Receipt.Status.SENT,
            provider='mock',
            payload={'receipt_number': 'R-1'},
            reprint_count=1,
        )
        self.shift_service.close_shift(
            shift=self.shift,
            actual_closing_cash_amount=130000,
            closed_by=self.user,
            notes_close='End of day',
        )

    def test_sales_report_is_scoped_to_branch(self):
        response = self.client.get('/api/v1/reports/sales/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.json()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['method'], Payment.Method.CASH)
        self.assertEqual(rows[0]['total'], self.closed_order.total)

    def test_open_checks_report_returns_current_branch_rows(self):
        response = self.client.get('/api/v1/reports/open-checks/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.json()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['orderNumber'], self.open_order.order_number)
        self.assertEqual(rows[0]['total'], self.open_order.total)

    def test_sales_report_export_returns_expected_columns(self):
        response = self.client.get('/api/v1/reports/sales/export/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        sales_columns = [str(label) for _key, label in get_sales_columns()]
        open_check_columns = [str(label) for key, label in get_open_checks_columns() if key not in {'hall_name', 'created_at'}]

        self.assertEqual(sheet['A1'].value, 'Ko\'rsatkich')
        self.assertEqual([sheet['A9'].value, sheet['B9'].value, sheet['C9'].value], sales_columns)
        self.assertEqual(sheet['A16'].value, 'Ochiq cheklar')
        self.assertEqual([sheet['A17'].value, sheet['B17'].value, sheet['C17'].value, sheet['D17'].value], open_check_columns)

    def test_admin_shift_report_returns_shift_rows(self):
        response = self.client.get('/api/v1/admin/reports/shifts/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['cash_desk_name'], self.cash_desk.name)
        self.assertEqual(rows[0]['cash_total'], self.closed_order.total)
        self.assertEqual(rows[0]['reprint_count'], 1)

    def test_admin_shift_report_export_returns_expected_columns(self):
        response = self.client.get('/api/v1/admin/reports/shifts/export/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        shift_columns = [str(label) for _key, label in get_shift_columns()]

        self.assertEqual(sheet['A4'].value, _('Period'))
        self.assertEqual([sheet.cell(row=7, column=index + 1).value for index in range(len(shift_columns))], shift_columns)
