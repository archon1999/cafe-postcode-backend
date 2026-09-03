from datetime import datetime, timedelta
from io import BytesIO

from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook
from rest_framework import status

from apps.sales.models import Order, OrderItem
from apps.billing.models import Payment, PaymentRefund, Receipt
from apps.billing.services import CashShiftService
from apps.reporting.services.export_localization import (
    REPORT_TITLE_SALES,
    REPORT_TITLE_SHIFTS,
    REPORT_TITLE_TOP_STAFF,
    get_report_title,
    get_report_export_filename,
    get_sales_columns,
    get_shift_columns,
    get_top_staff_columns,
)
from apps.reporting.selectors.reporting import get_report_period
from apps.reporting.services.report_excel_export import ReportExcelExportService
from apps.sales.tests.support.pos_api import PosAPITestCase
from common.utils.date import TASHKENT_TIMEZONE


class ReportsApiTests(PosAPITestCase):
    def test_excel_export_escapes_formula_prefixes(self):
        payload = ReportExcelExportService().build_table_file(
            title='Security export',
            columns=(('name', 'Name'),),
            rows=({'name': '=HYPERLINK("https://example.invalid", "click")'},),
        )

        sheet = load_workbook(filename=BytesIO(payload)).active
        cell = sheet['A4']
        self.assertEqual(cell.data_type, 's')
        self.assertTrue(cell.value.startswith("'="))

    def setUp(self):
        super().setUp()
        now = timezone.localtime(timezone.now(), TASHKENT_TIMEZONE).replace(
            hour=12,
            minute=0,
            second=0,
            microsecond=0,
        )
        self.report_date = now.date()
        self.shift_service = CashShiftService()
        self.closed_order = Order.objects.create(
            restaurant=self.restaurant,
            distribution_point=self.takeaway_distribution,
            opened_by=self.user,
            cashier=self.user,
            order_number=1,
            channel=Order.Channel.TAKEAWAY,
            status=Order.Status.CLOSED,
            guest_count=1,
            closed_at=now - timedelta(minutes=15),
        )
        OrderItem.objects.create(
            order=self.closed_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=30000,
            status=OrderItem.Status.DONE,
        )
        self.closed_order.recalculate_totals()
        second_restaurant = self.restaurant.__class__.objects.create(
            name='Other restaurant',
            service_fee_percent=10,
        )
        other_order = Order.objects.create(
            restaurant=second_restaurant,
            distribution_point=self.takeaway_distribution.__class__.objects.create(
                restaurant=second_restaurant,
                name='Other takeaway',
                kind=self.takeaway_distribution.kind,
            ),
            opened_by=self.user,
            cashier=self.user,
            order_number=1,
            channel=Order.Channel.TAKEAWAY,
            status=Order.Status.CLOSED,
            guest_count=1,
            closed_at=now - timedelta(hours=1),
        )
        OrderItem.objects.create(
            order=other_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=50000,
            status=OrderItem.Status.DONE,
        )
        other_order.recalculate_totals()
        Payment.objects.create(
            order=other_order,
            received_by=self.user,
            method=Payment.Method.CASH,
            amount=other_order.total,
            status=Payment.Status.SUCCEEDED,
            paid_at=now - timedelta(hours=1),
        )

        self.shift = self.create_cash_shift(opening_cash_amount=100000)
        self.shift_payment = Payment.objects.create(
            order=self.closed_order,
            cash_shift=self.shift,
            cash_desk=self.cash_desk,
            received_by=self.user,
            method=Payment.Method.CASH,
            amount=self.closed_order.total,
            status=Payment.Status.SUCCEEDED,
            paid_at=now - timedelta(minutes=30),
        )
        Receipt.objects.create(
            order=self.closed_order,
            payment=self.shift_payment,
            kind=Receipt.Kind.FISCAL,
            status=Receipt.Status.SENT,
            provider='mock',
            payload={'receipt_number': 'R-1'},
            reprint_count=1,
        )
        Receipt.objects.create(
            order=self.closed_order,
            payment=self.shift_payment,
            kind=Receipt.Kind.PLAIN,
            status=Receipt.Status.SENT,
            provider='mock',
            payload={'receipt_number': 'P-1'},
        )
        self.shift_service.close_shift(
            shift=self.shift,
            actual_closing_cash_amount=130000,
            closed_by=self.user,
            notes_close='End of day',
        )

        # Reports still need an open order in the restaurant, but it must be
        # created after the historical shift is closed. Production shift close
        # now correctly rejects restaurants with unresolved open orders.
        self.open_order = Order.objects.create(
            restaurant=self.restaurant,
            distribution_point=self.hall_distribution,
            opened_by=self.user,
            order_number=2,
            channel=Order.Channel.HALL,
            status=Order.Status.SUBMITTED,
            guest_count=2,
        )
        OrderItem.objects.create(
            order=self.open_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=1,
            unit_price=30000,
            status=OrderItem.Status.NEW,
        )
        self.open_order.recalculate_totals()
        OrderItem.objects.create(
            order=self.open_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            created_by=self.user,
            quantity=3,
            unit_price=10000,
            status=OrderItem.Status.CANCELLED,
        )

    def current_range_params(self):
        current_date = self.report_date.isoformat()
        return {
            'start_date': current_date,
            'end_date': current_date,
        }

    def test_sales_report_is_scoped_to_branch(self):
        response = self.client.get('/api/v1/admin/reporting/sales/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['method'], Payment.Method.CASH)
        self.assertEqual(rows[0]['total'], self.closed_order.total)

    def test_sales_report_uses_tashkent_midnight_boundary(self):
        report_date = datetime(2031, 1, 15, tzinfo=TASHKENT_TIMEZONE)
        Payment.objects.create(
            order=self.closed_order,
            received_by=self.user,
            method=Payment.Method.QR,
            amount=11000,
            status=Payment.Status.SUCCEEDED,
            paid_at=report_date - timedelta(minutes=1),
        )
        Payment.objects.create(
            order=self.closed_order,
            received_by=self.user,
            method=Payment.Method.CARD,
            amount=22000,
            status=Payment.Status.SUCCEEDED,
            paid_at=report_date + timedelta(minutes=1),
        )

        response = self.client.get(
            '/api/v1/admin/reporting/sales/',
            {
                'start_date': report_date.date().isoformat(),
                'end_date': report_date.date().isoformat(),
            },
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data['data'],
            [{'method': Payment.Method.CARD, 'count': 1, 'total': 22000}],
        )

    def test_summary_separates_gross_refunds_and_net_sales(self):
        PaymentRefund.objects.create(
            payment=self.shift_payment,
            amount=5000,
            refunded_by=self.user,
            status=PaymentRefund.Status.SUCCEEDED,
            refunded_at=timezone.now(),
        )

        response = self.client.get('/api/v1/admin/reporting/summary/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['gross_sales_total'], self.closed_order.total)
        self.assertEqual(response.data['refunds_total'], 5000)
        self.assertEqual(response.data['sales_total'], self.closed_order.total - 5000)
        self.assertEqual(response.data['average_check'], self.closed_order.total - 5000)
        self.assertEqual(response.data['prechecks_count'], 1)
        self.assertEqual(response.data['receipts_count'], 1)
        self.assertNotIn('open_checks', response.data)
        self.assertNotIn('active_tables', response.data)

    def test_open_checks_report_returns_current_branch_rows(self):
        response = self.client.get('/api/v1/admin/reporting/open-checks/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['order_number'], self.open_order.order_number)
        self.assertEqual(rows[0]['total'], self.open_order.total)

    def test_summary_export_contains_precheck_and_receipt_counts(self):
        response = self.client.get('/api/v1/admin/reporting/summary/export/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        sheet = load_workbook(filename=BytesIO(response.content)).active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn(_('Prechecks'), values)
        self.assertIn(_('Receipts'), values)
        self.assertNotIn(_('Open Checks'), values)
        self.assertNotIn(_('Active Tables'), values)

    def test_receipts_report_returns_prechecks_and_receipts(self):
        response = self.client.get('/api/v1/admin/reporting/receipts/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 2)
        self.assertEqual({row['kind'] for row in rows}, {Receipt.Kind.PLAIN, Receipt.Kind.FISCAL})
        self.assertTrue(all(row['order_number'] == self.closed_order.order_number for row in rows))
        self.assertTrue(all(row['amount'] == self.shift_payment.amount for row in rows))

    def test_receipts_report_filters_by_kind(self):
        params = {**self.current_range_params(), 'receipt_kind': Receipt.Kind.PLAIN}
        response = self.client.get('/api/v1/admin/reporting/receipts/', params)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['data'][0]['kind'], Receipt.Kind.PLAIN)

    def test_receipts_report_export_is_localized(self):
        response = self.client.get('/api/v1/admin/reporting/receipts/export/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        sheet = load_workbook(filename=BytesIO(response.content)).active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn(_('Precheck'), values)
        self.assertIn(_('Receipt'), values)

    def test_sales_report_export_returns_expected_columns(self):
        range_params = self.current_range_params()
        response = self.client.get('/api/v1/admin/reporting/sales/export/', range_params)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        expected_filename = get_report_export_filename(REPORT_TITLE_SALES, get_report_period(response.wsgi_request.GET))
        self.assertIn(expected_filename, response['Content-Disposition'])
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        sales_columns = [str(label) for _key, label in get_sales_columns()]

        self.assertEqual(sheet['A1'].value, str(get_report_title(REPORT_TITLE_SALES)))
        self.assertEqual(sheet['A3'].value, _('Filter'))
        self.assertEqual([sheet['A7'].value, sheet['B7'].value, sheet['C7'].value], sales_columns)

    def test_admin_shift_report_returns_shift_rows(self):
        response = self.client.get('/api/v1/admin/reporting/shifts/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['cash_desk_name'], self.cash_desk.name)
        self.assertEqual(rows[0]['cash_total'], self.closed_order.total)
        self.assertEqual(rows[0]['precheck_count'], 1)
        self.assertEqual(rows[0]['reprint_count'], 1)

    def test_top_staff_report_uses_order_item_creator_totals(self):
        response = self.client.get('/api/v1/admin/reporting/top-staff/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['staff_name'], self.user.full_name)
        self.assertEqual(rows[0]['order_count'], 1)
        self.assertEqual(rows[0]['items_count'], 1)
        self.assertEqual(rows[0]['total_sales'], self.closed_order.subtotal)

    def test_top_items_report_only_uses_closed_orders(self):
        response = self.client.get('/api/v1/admin/reporting/top-items/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        rows = response.data['data']
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['quantity'], 1)
        self.assertEqual(rows[0]['sale_unit'], 'piece')
        self.assertEqual(rows[0]['item_type'], 'product')
        self.assertEqual(rows[0]['revenue'], self.closed_order.subtotal)

    def test_top_staff_report_labels_unknown_item_creator(self):
        unknown_order = Order.objects.create(
            restaurant=self.restaurant,
            distribution_point=self.takeaway_distribution,
            order_number=3,
            channel=Order.Channel.TAKEAWAY,
            status=Order.Status.CLOSED,
            guest_count=1,
            closed_at=timezone.now(),
        )
        OrderItem.objects.create(
            order=unknown_order,
            catalog_item=self.catalog_item,
            prep_station=self.prep_station,
            quantity=1,
            unit_price=15000,
            status=OrderItem.Status.NEW,
        )

        response = self.client.get('/api/v1/admin/reporting/top-staff/', self.current_range_params())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        unknown_row = next(row for row in response.data['data'] if row['staff_id'] is None)
        self.assertEqual(unknown_row['staff_name'], "Noma'lum")
        self.assertEqual(unknown_row['items_count'], 1)

    def test_admin_shift_report_export_returns_expected_columns(self):
        range_params = self.current_range_params()
        response = self.client.get('/api/v1/admin/reporting/shifts/export/', range_params)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        expected_filename = get_report_export_filename(REPORT_TITLE_SHIFTS, get_report_period(response.wsgi_request.GET))
        self.assertIn(expected_filename, response['Content-Disposition'])
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        shift_columns = [str(label) for _key, label in get_shift_columns()]
        shift_column_keys = [key for key, _label in get_shift_columns()]

        self.assertEqual(sheet['A1'].value, str(get_report_title(REPORT_TITLE_SHIFTS)))
        self.assertEqual(sheet['A4'].value, _('Period'))
        self.assertEqual([sheet.cell(row=7, column=index + 1).value for index in range(len(shift_columns))], shift_columns)
        self.assertIn('precheck_count', shift_column_keys)
        self.assertNotIn('qr_total', shift_column_keys)
        self.assertNotIn('reprint_count', shift_column_keys)
        self.assertEqual(shift_column_keys[-1], 'id')

    def test_shift_export_localizes_status_and_filter_values(self):
        params = {**self.current_range_params(), 'status': 'closed', 'difference_only': 'true'}
        response = self.client.get('/api/v1/admin/reporting/shifts/export/', params)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        sheet = load_workbook(filename=BytesIO(response.content)).active
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
        self.assertIn(_('Closed'), values)
        self.assertIn(_('Yes'), values)

    def test_top_staff_report_export_returns_items_column(self):
        range_params = self.current_range_params()
        response = self.client.get('/api/v1/admin/reporting/top-staff/export/', range_params)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        top_staff_columns = [str(label) for _key, label in get_top_staff_columns()]

        self.assertEqual(sheet['A1'].value, str(get_report_title(REPORT_TITLE_TOP_STAFF)))
        self.assertEqual([sheet.cell(row=7, column=index + 1).value for index in range(len(top_staff_columns))], top_staff_columns)

