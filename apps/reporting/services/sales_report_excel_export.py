from django.utils.translation import gettext as _
from openpyxl import Workbook

from .export_localization import (
    REPORT_TITLE_OPEN_CHECKS,
    REPORT_TITLE_PAYMENT_BREAKDOWN,
    REPORT_TITLE_SALES,
    get_open_checks_columns,
    get_report_section_title,
    get_sales_columns,
    get_summary_metrics,
    localize_open_checks_rows,
    localize_payment_breakdown_rows,
    localize_sales_rows,
)
from .report_excel_export import ReportExcelExportService


class SalesReportExcelExportService:
    sheet_title = _('Daily Sales')

    def build_file(self, summary: dict, sales_rows: list[dict], payment_rows: list[dict], open_check_rows: list[dict]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_title

        helper = ReportExcelExportService()
        sheet.append([_('Metric'), _('Value')])
        helper._make_bold(sheet[1])

        for label, value in get_summary_metrics(summary):
            sheet.append([label, helper._normalize_value(value)])
        sheet.append([])

        section_start_row = sheet.max_row + 1
        sheet.append([get_report_section_title(REPORT_TITLE_SALES), '', ''])
        helper._make_bold(sheet[section_start_row])
        sheet.append([label for _key, label in get_sales_columns()])
        helper._make_bold(sheet[sheet.max_row])
        for row in localize_sales_rows(sales_rows):
            sheet.append([helper._normalize_value(row.get(key)) for key, _label in get_sales_columns()])

        sheet.append([])
        section_start_row = sheet.max_row + 1
        sheet.append([get_report_section_title(REPORT_TITLE_PAYMENT_BREAKDOWN), '', ''])
        helper._make_bold(sheet[section_start_row])
        sheet.append([label for _key, label in get_sales_columns()])
        helper._make_bold(sheet[sheet.max_row])
        for row in localize_payment_breakdown_rows(payment_rows):
            sheet.append([helper._normalize_value(row.get(key)) for key, _label in get_sales_columns()])

        sheet.append([])
        section_start_row = sheet.max_row + 1
        sheet.append([get_report_section_title(REPORT_TITLE_OPEN_CHECKS), '', '', ''])
        helper._make_bold(sheet[section_start_row])
        sheet.append([label for _key, label in get_open_checks_columns() if _key != 'hall_name' and _key != 'created_at'])
        helper._make_bold(sheet[sheet.max_row])
        for row in localize_open_checks_rows(open_check_rows):
            sheet.append(
                [
                    helper._normalize_value(row.get('order_number')),
                    helper._normalize_value(row.get('table_name')),
                    helper._normalize_value(row.get('status')),
                    helper._normalize_value(row.get('total')),
                ]
            )

        for column in ['A', 'B', 'C', 'D']:
            sheet.column_dimensions[column].width = 24

        return helper._dump(workbook)
