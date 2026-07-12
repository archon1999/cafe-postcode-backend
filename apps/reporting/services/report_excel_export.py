import json
from collections.abc import Mapping, Sequence
from datetime import datetime
from io import BytesIO
from uuid import UUID

from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font

from common.utils.date import localize_to_tashkent

from .export_localization import get_empty_value_placeholder


class ReportExcelExportService:
    def build_summary_file(
        self,
        *,
        title: str,
        metrics: Sequence[tuple[str, object]],
        filters: Sequence[tuple[str, str]] = (),
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._sheet_title(title)
        self._write_header(sheet, title)
        self._write_filters(sheet, filters)
        sheet.append([_('Metric'), _('Value')])
        self._make_bold(sheet[sheet.max_row])
        for label, value in metrics:
            sheet.append([label, self._normalize_value(value)])
        self._apply_widths(sheet)
        return self._dump(workbook)

    def build_table_file(
        self,
        *,
        title: str,
        columns: Sequence[tuple[str, str]],
        rows: Sequence[Mapping[str, object]],
        filters: Sequence[tuple[str, str]] = (),
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._sheet_title(title)
        self._write_header(sheet, title)
        self._write_filters(sheet, filters)
        sheet.append([label for _key, label in columns])
        self._make_bold(sheet[sheet.max_row])
        for row in rows:
            sheet.append([self._normalize_value(row.get(key)) for key, _label in columns])
        self._apply_widths(sheet)
        return self._dump(workbook)

    def _write_header(self, sheet, title: str) -> None:
        sheet.append([title])
        self._make_bold(sheet[1], size=14)
        sheet.append([])

    def _write_filters(self, sheet, filters: Sequence[tuple[str, str]]) -> None:
        if not filters:
            return
        sheet.append([_('Filter'), _('Value')])
        self._make_bold(sheet[sheet.max_row])
        for label, value in filters:
            sheet.append([self._normalize_value(label), self._normalize_value(value)])
        sheet.append([])

    def _apply_widths(self, sheet) -> None:
        for column_cells in sheet.columns:
            values = [len(str(cell.value or '')) for cell in column_cells]
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max(values + [12]) + 2, 36)

    @staticmethod
    def _sheet_title(title: str) -> str:
        return title[:31]

    def _normalize_value(self, value: object) -> object:
        if value is None or value == '':
            return get_empty_value_placeholder()
        if isinstance(value, datetime):
            value = localize_to_tashkent(value)
            return value.replace(tzinfo=None)
        if isinstance(value, UUID):
            return str(value)
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False)
        if isinstance(value, str) and value.startswith(('=', '+', '-', '@', '\t', '\r')):
            return "'" + value
        return value

    @staticmethod
    def _make_bold(cells, *, size: int = 11) -> None:
        font = Font(bold=True, size=size)
        for cell in cells:
            cell.font = font

    @staticmethod
    def _dump(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
